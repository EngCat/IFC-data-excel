# Modified IFC Extractor with GUID in first column
import ifcopenshell
import pandas as pd
import tkinter as tk
from tkinter import messagebox
from tkinter.filedialog import askopenfilename, asksaveasfilename
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def main():
    def on_submit():
        if not ifc_file_path.get():
            messagebox.showwarning("Warning", "You must select an IFC file!")
            return
        if not excel_file_path.get():
            messagebox.showwarning("Warning", "You must select a save location for the Excel file!")
            return
        selected_categories = [category_listbox.get(i) for i in category_listbox.curselection()]
        if not selected_categories:
            messagebox.showwarning("Warning", "You must select at least one category!")
            return
        run_combined_extraction(selected_categories, ifc_file_path.get(), excel_file_path.get())
        root.destroy()

    def browse_ifc_file():
        file_path = askopenfilename(title="Select the IFC file", filetypes=[("IFC files", "*.ifc")])
        if file_path:
            ifc_file_path.set(file_path)
            label_ifc_file.config(text=f"IFC File: {file_path}")
            load_categories(file_path)

    def browse_excel_file():
        file_path = asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], title="Save Excel file as")
        if file_path:
            excel_file_path.set(file_path)
            label_excel_file.config(text=f"Excel File: {file_path}")

    def load_categories(ifc_file_path):
        all_categories = load_ifc_data(ifc_file_path)
        update_category_options(all_categories)

    def load_ifc_data(ifc_file_path):
        ifc_file = load_ifc_file(ifc_file_path)
        if ifc_file is None:
            return []
        categories = sorted({entity.is_a() for entity in ifc_file.by_type("IfcProduct")})
        return categories

    def update_category_options(categories):
        category_listbox.delete(0, tk.END)
        for category in categories:
            category_listbox.insert(tk.END, category)

    def on_select_all():
        category_listbox.select_set(0, tk.END)

    def on_deselect_all():
        category_listbox.selection_clear(0, tk.END)

    root = tk.Tk()
    root.title("IFC Extractor")
    border_color = '#5D8A66'

    ifc_file_path = tk.StringVar()
    excel_file_path = tk.StringVar()

    tk.Button(root, text="Browse IFC File", bg="white", fg=border_color, command=browse_ifc_file).pack(pady=5)
    label_ifc_file = tk.Label(root, text="IFC File: Not selected", bg='#5D8A66', fg='white')
    label_ifc_file.pack(pady=5)

    tk.Button(root, text="Browse Excel File", bg="white", fg=border_color, command=browse_excel_file).pack(pady=5)
    label_excel_file = tk.Label(root, text="Excel File: Not selected", bg='#5D8A66', fg='white')
    label_excel_file.pack(pady=5)

    tk.Label(root, text="Select categories to include:", bg='#5D8A66', fg='white').pack(pady=10)
    category_listbox = tk.Listbox(root, selectmode=tk.MULTIPLE)
    category_listbox.pack(pady=10, fill=tk.BOTH, expand=True)

    tk.Button(root, text="Select All Categories", bg="white", fg=border_color, command=on_select_all).pack(pady=5)
    tk.Button(root, text="Deselect All Categories", bg="white", fg=border_color, command=on_deselect_all).pack(pady=5)

    tk.Button(root, text="Extract", bg="white", fg=border_color, command=on_submit).pack(pady=10)

    root.geometry("400x600")
    root.configure(bg='#5D8A66')
    root.mainloop()


def load_ifc_file(path):
    try:
        return ifcopenshell.open(path)
    except Exception as e:
        print("Error loading IFC:", e)
        return None


def run_combined_extraction(categories, ifc_path, excel_path):
    ifc_file = load_ifc_file(ifc_path)
    if ifc_file is None:
        return

    writer = pd.ExcelWriter(excel_path, engine='openpyxl')

    for category in categories:
        data = extract_category_data(ifc_file, category)
        if data:
            df = pd.DataFrame(data)
            df.to_excel(writer, sheet_name=f"{category}_Properties", index=False)

    writer.close()
    customize_excel(excel_path)
    messagebox.showinfo("Done", f"Saved: {excel_path}")


def extract_category_data(ifc_file, category):
    out = []
    for element in ifc_file.by_type(category):
        element_data = {
            "GUID": element.GlobalId,   # GUID FIRST COLUMN
            "Category": category
        }

        if hasattr(element, "Name"):
            element_data["Name"] = element.Name

        # PROPERTY SET EXTRACTION
        for rel in getattr(element, "IsDefinedBy", []):
            if rel.is_a("IfcRelDefinesByProperties"):
                prop_set = rel.RelatingPropertyDefinition
                if prop_set.is_a("IfcPropertySet"):
                    for prop in prop_set.HasProperties:
                        if prop.is_a("IfcPropertySingleValue"):
                            value = prop.NominalValue
                            if value:
                                element_data[f"{prop_set.Name} - {prop.Name}"] = value.wrappedValue
                        elif prop.is_a("IfcPropertyEnumeratedValue"):
                            vals = prop.EnumerationValues
                            if vals:
                                element_data[f"{prop_set.Name} - {prop.Name}"] = [v.wrappedValue for v in vals]

        out.append(element_data)

    return out


def customize_excel(excel_path):
    wb = load_workbook(excel_path)
    fill = PatternFill(start_color="5D8A66", end_color="5D8A66", fill_type="solid")

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for cell in ws[1]:
            cell.fill = fill
        for col in ws.columns:
            maxlen = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = max(maxlen, 10)

    wb.save(excel_path)


if __name__ == "__main__":
    main()
